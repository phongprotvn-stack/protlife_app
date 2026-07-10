#!/usr/bin/env python3
"""Import Prot_LifeOS_v1.0.0.xlsm → JSON for web app import."""
import openpyxl, json, os, sys, re
from datetime import datetime

EXCEL_PATH = r'D:\CODE\Database\ProtLifeApp\Prot_LifeOS_v1.0.0.xlsm'
OUTPUT_JSON = r'D:\CODE\P_projects\protlife_app\scripts\protlife_import_data.json'
OUTPUT_FIRESTORE = r'D:\CODE\P_projects\protlife_app\scripts\firestore_import.json'

STATUS_MAP = {'Active': 'Active', 'Lost Contact': 'Lost Contact', 'Deceased': 'Deceased', 'Blocked': 'Blocked'}
GENDER_MAP = {'Male': 'male', 'Female': 'female', 'Other': 'other'}
BOOL_MAP = {'TRUE': True, 'FALSE': False, True: True, False: False}

def parse_date(v):
    """Convert various date formats to YYYY-MM-DD or dd/mm/yyyy."""
    if not v or str(v).strip() == '':
        return ''
    v = str(v).strip()
    # Already YYYY-MM-DD
    if re.match(r'^\d{4}-\d{2}-\d{2}', v):
        return v[:10]
    # Excel datetime serial
    if isinstance(v, (int, float)) and v > 40000:
        try:
            from datetime import datetime as dt
            d = dt.fromordinal(dt(1900, 1, 1).toordinal() + int(v) - 2)
            return d.strftime('%Y-%m-%d')
        except:
            pass
    # dd/mm/yyyy or d/m/yyyy
    m = re.search(r'(\d{1,2})[/-](\d{1,2})[/-](\d{4})', v)
    if m:
        return f'{m.group(1).zfill(2)}/{m.group(2).zfill(2)}/{m.group(3)}'
    return v

def parse_gender(v):
    if not v:
        return 'male'
    v = str(v).strip().lower()
    if 'female' in v or 'nữ' in v or v == 'f':
        return 'female'
    if 'male' in v or 'nam' in v or v == 'm':
        return 'male'
    return 'other'

def make_id(prefix, num):
    return f'{prefix}{str(num).zfill(5)}'

print("📖 Reading Excel file...")
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, keep_vba=False)

# === READ CONTACTS ===
ws = wb['Contact']
contacts = []
for row in ws.iter_rows(min_row=3, values_only=True):
    if row[0] is None or str(row[0]).strip() == '':
        continue
    cid, name, rel, bday, gender, phone, email, org0, org1, org2, score, status, created, updated, fav = row[:15]
    
    # Build organizations from 3 org fields
    orgs = []
    for o in [org0, org1, org2]:
        if o and str(o).strip():
            orgs.append(str(o).strip())
    
    now = datetime.now().isoformat()
    person = {
        'id': make_id('P', contacts.__len__() + 1),
        'legacyId': str(cid).strip() if cid else '',
        'name': str(name).strip() if name else '',
        'relationship': str(rel).strip() if rel else '',
        'dob': parse_date(bday),
        'gender': parse_gender(gender),
        'phones': [str(phone).strip()] if phone and str(phone).strip() else [],
        'emails': [str(email).strip()] if email and str(email).strip() else [],
        'organizations': orgs,
        'organization': orgs[0] if orgs else '',
        'relationshipScore': int(score) if score else 50,
        'status': STATUS_MAP.get(str(status).strip(), 'Active') if status else 'Active',
        'isFavorite': BOOL_MAP.get(fav, False) if fav is not None else False,
        'source': 'Excel Import',
        'notes': '',
        'socialLinks': [],
        'groupIds': [],
        'tags': [],
        'address': '',
        'createdAt': parse_date(created) or now,
        'updatedAt': parse_date(updated) or now,
    }
    contacts.append(person)
    if len(contacts) <= 3:
        print(f"  Sample: {person['name']} | rel={person['relationship']} | score={person['relationshipScore']} | status={person['status']} | fav={person['isFavorite']} | orgs={orgs}")

print(f"\n✅ Total contacts: {len(contacts)}")

# === READ EVENTS ===
ws2 = wb['Event']
COST_MAP = {'': 0, None: 0}
events = []
for row in ws2.iter_rows(min_row=3, values_only=True):
    if row[0] is None or str(row[0]).strip() == '':
        continue
    eid, no, title, participants, etype, stage, source, start_date, end_date, place, maplink, mood, importance, pcount, cost, notes, created, updated = row[:18]
    
    if not title or str(title).strip() == '':
        continue
    
    now = datetime.now().isoformat()
    event = {
        'id': make_id('EV', events.__len__() + 1),
        'legacyId': str(eid).strip() if eid else '',
        'title': str(title).strip(),
        'eventType': str(etype).strip().lower() if etype else '',
        'date': parse_date(start_date),
        'endDate': parse_date(end_date),
        'mood': str(mood).strip() if mood else '',
        'importance': str(importance).strip() if importance else '',
        'lifeStage': str(stage).strip() if stage else '',
        'source': str(source).strip() if source else '',
        'cost': int(cost) if cost and str(cost).strip().isdigit() else 0,
        'mapLink': str(maplink).strip() if maplink and str(maplink).strip() != 'Map' else '',
        'locationName': str(place).strip() if place else '',
        'peopleIds': [],
        'participants': str(participants).strip() if participants else '',
        'notes': str(notes).strip() if notes else '',
        'participantCount': int(pcount) if pcount and str(pcount).strip().isdigit() else 0,
        'createdAt': parse_date(created) or now,
        'updatedAt': parse_date(updated) or now,
    }
    events.append(event)

print(f"✅ Total events: {len(events)}")

# === READ EVENT PARTICIPANTS ===
ws3 = wb['EventParticipant']
participants = []
for row in ws3.iter_rows(min_row=3, values_only=True):
    if row[0] is None or str(row[0]).strip() == '':
        continue
    rec_id, event_id, contact_id, contact_name, role, attendance = row[:6]
    participants.append({
        'recordId': str(rec_id).strip() if rec_id else '',
        'eventLegacyId': str(event_id).strip() if event_id else '',
        'contactLegacyId': str(contact_id).strip() if contact_id else '',
        'contactName': str(contact_name).strip() if contact_name else '',
        'role': str(role).strip() if role else 'Participant',
        'attendance': str(attendance).strip() if attendance else 'Confirmed',
    })

print(f"✅ Total event-participant links: {len(participants)}")

# === BUILD LOOKUP ===
lookup = {}
current_cat = None
ws4 = wb['Lookup']
for row in ws4.iter_rows(min_row=1, values_only=True):
    val = row[0]
    if val is None or str(val).strip() == '':
        continue
    v = str(val).strip()
    if v == v.upper() and len(v) > 3 and not v.startswith('LOOKUP'):
        current_cat = v
        lookup[current_cat] = []
    elif current_cat and v:
        lookup[current_cat].append(v)

print(f"✅ Lookup categories: {list(lookup.keys())}")

# === MATCH PARTICIPANTS TO CONTACT IDS ===
# Build legacyId → newId mapping
legacy_map = {}
for p in contacts:
    if p['legacyId']:
        legacy_map[p['legacyId']] = p['id']

# Also try to match by name
name_map = {}
for p in contacts:
    name_key = p['name'].strip().lower().replace(' ', '')
    name_map[name_key] = p['id']

# Match event participants
for ep in participants:
    matched_id = None
    # Try by contact legacyId
    if ep['contactLegacyId'] and ep['contactLegacyId'] in legacy_map:
        matched_id = legacy_map[ep['contactLegacyId']]
    # Try by name
    if not matched_id:
        name_key = ep['contactName'].strip().lower().replace(' ', '')
        matched_id = name_map.get(name_key)
    
    if matched_id:
        # Find the event by legacyId and add the person ID
        for ev in events:
            if ev['legacyId'] == ep['eventLegacyId']:
                if matched_id not in ev['peopleIds']:
                    ev['peopleIds'].append(matched_id)
                break

print("✅ Participants matched to people IDs")

# === GENERATE OUTPUT JSON ===
output = {
    'people': contacts,
    'events': events,
    'memories': [],
    'lookup': lookup,
    'participants': participants,
    '_meta': {
        'source': 'Prot_LifeOS_v1.0.0.xlsm',
        'exportedAt': datetime.now().isoformat(),
        'totalContacts': len(contacts),
        'totalEvents': len(events),
        'totalParticipants': len(participants),
    }
}

os.makedirs(os.path.dirname(OUTPUT_JSON) or '.', exist_ok=True)
with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print(f"\n📝 Output written to:")
print(f"   {OUTPUT_JSON}")
print(f"   Size: {os.path.getsize(OUTPUT_JSON):,} bytes")

# === GENERATE FIRESTORE IMPORT JSON ===
# Structure: one doc per collection
firestore_data = {
    'users/phongprot.vn@gmail.com/people': {p['id']: p for p in contacts},
    'users/phongprot.vn@gmail.com/events': {e['id']: e for e in events},
}

with open(OUTPUT_FIRESTORE, 'w', encoding='utf-8') as f:
    json.dump(firestore_data, f, ensure_ascii=False, indent=2)

print(f"   {OUTPUT_FIRESTORE}")
print(f"   Size: {os.path.getsize(OUTPUT_FIRESTORE):,} bytes")
print("\n🎉 Done!")
